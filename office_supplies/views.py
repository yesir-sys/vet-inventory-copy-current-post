from itertools import count
from tkinter.font import Font
from unicodedata import category
from django.forms import ValidationError
from django.views import View
import csv
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from io import TextIOWrapper
from django.utils import timezone
from django.urls import reverse_lazy
from django.views.generic import ListView, CreateView, UpdateView, DeleteView, DetailView, TemplateView
from django.contrib.auth.views import LoginView, LogoutView
from django.contrib.auth.forms import UserCreationForm
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth import login
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.db import transaction, IntegrityError
from django.db.models import Q, F, Count, Sum

from .models import OfficeSupply, OfficeCategory, OfficeMassOutgoing
from .forms import OfficeSupplyForm, OfficeMassOutgoingForm, OfficeMassOutgoingItemFormSet, MassAddForm
from django.http import HttpResponse, JsonResponse
from datetime import datetime, timedelta, date
from django.contrib import messages
from django.shortcuts import redirect
from django.views.generic import CreateView
from django.urls import reverse_lazy
from .models import OfficeMassOutgoing
from .forms import OfficeMassOutgoingForm, OfficeMassOutgoingItemFormSet
from django.db import transaction
from django.core.exceptions import ValidationError

import io
from django.contrib.auth.decorators import login_required
from .models import OfficeMassIncoming
from django.contrib.contenttypes.models import ContentType
from reports.models import InventoryMovement

from core.mixins import AdminRequiredMixin

# Authentication Views
class UserSignUpView(CreateView):
    form_class = UserCreationForm
    template_name = 'registration/signup.html'
    success_url = reverse_lazy('office_supplies:supply-list')  # Updated
    
    def form_valid(self, form):
        response = super().form_valid(form)
        login(self.request, self.object)
        return response

class UserLoginView(LoginView):
    template_name = 'registration/login.html'

class UserLogoutView(LogoutView):
    next_page = reverse_lazy('office_supplies:supply-list')  # Updated

# Office Supply Views
class OfficeSupplyList(LoginRequiredMixin, ListView):
    model = OfficeSupply
    template_name = 'office_supplies/list.html'
    context_object_name = 'supplies'
    paginate_by = 10

    def get_queryset(self):
        queryset = OfficeSupply.objects.all().order_by('name')
        
        # Search query
        search_query = self.request.GET.get('q')
        if (search_query):
            queryset = queryset.filter(
                Q(name__icontains=search_query) |
                Q(category__name__icontains=search_query) |
                Q(description__icontains=search_query)
            )

        # Category filter
        category = self.request.GET.get('category')
        if (category):
            queryset = queryset.filter(category_id=category)

        # Stock status filter
        stock_status = self.request.GET.get('stock_status')
        if (stock_status == 'low'):
            queryset = queryset.filter(quantity__lte=F('reorder_level'))
        elif (stock_status == 'normal'):
            queryset = queryset.filter(quantity__gt=F('reorder_level'))

        # Expiration status filter
        today = timezone.now().date()
        thirty_days_later = today + timedelta(days=30)
        expiration_status = self.request.GET.get('expiration_status')
        
        if (expiration_status == 'expired'):
            queryset = queryset.filter(expiration_date__lt=today)
        elif (expiration_status == 'expiring_soon'):
            queryset = queryset.filter(
                expiration_date__gte=today,
                expiration_date__lte=thirty_days_later
            )
        elif (expiration_status == 'not_expiring'):
            queryset = queryset.filter(
                Q(expiration_date__gt=thirty_days_later) |
                Q(expiration_date__isnull=True)
            )

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        today = timezone.now().date()
        thirty_days_later = today + timedelta(days=30)
        
        context.update({
            'categories': OfficeCategory.objects.all(),
            'selected_category': self.request.GET.get('category', ''),
            'search_query': self.request.GET.get('q', ''),
            'stock_status': self.request.GET.get('stock_status', ''),
            'expiration_status': self.request.GET.get('expiration_status', ''),
            'low_stock_count': OfficeSupply.objects.filter(
                quantity__lte=F('reorder_level')
            ).count(),
            'expiring_soon_count': OfficeSupply.objects.filter(
                expiration_date__gt=today,
                expiration_date__lte=thirty_days_later
            ).count(),
        })
        
        # Preserve query parameters in pagination
        if (self.request.GET):
            query_params = self.request.GET.copy()
            if ('page' in query_params):
                del query_params['page']
            context['query_params'] = query_params.urlencode()
        
        return context

def get_supplies_json(request):
    supplies = OfficeSupply.objects.all().values('name', 'quantity', 'reorder_level', 'expiration_date', 'expiration_status')
    return JsonResponse(list(supplies), safe=False)

class OfficeSupplyCreate(AdminRequiredMixin, LoginRequiredMixin, CreateView):
    model = OfficeSupply
    form_class = OfficeSupplyForm
    template_name = 'office_supplies/form.html'
    
    def get_success_url(self):
        return reverse_lazy('office_supplies:supply-list')

class OfficeSupplyUpdate(AdminRequiredMixin, LoginRequiredMixin, UpdateView):
    model = OfficeSupply
    form_class = OfficeSupplyForm
    template_name = 'office_supplies/form.html'
    
    def get_success_url(self):
        return reverse_lazy('office_supplies:supply-list')

class OfficeSupplyDelete(AdminRequiredMixin, LoginRequiredMixin, DeleteView):
    model = OfficeSupply
    template_name = 'office_supplies/confirm_delete.html'
    success_url = reverse_lazy('office_supplies:supply-list')

# Category Views
class OfficeCategoryList(LoginRequiredMixin, ListView):
    model = OfficeCategory
    template_name = 'office_supplies/category_list.html'
    context_object_name = 'categories'
    ordering = ['name']

class OfficeCategoryCreate(AdminRequiredMixin, LoginRequiredMixin, CreateView):
    model = OfficeCategory
    template_name = 'office_supplies/category_form.html'
    fields = ['name', 'description']
    success_url = reverse_lazy('office_supplies:category-list')

class OfficeCategoryUpdate(AdminRequiredMixin, LoginRequiredMixin, UpdateView):
    model = OfficeCategory
    fields = ['name']
    template_name = 'office_supplies/category_form.html'
    success_url = reverse_lazy('office_supplies:category-list')

class OfficeCategoryDelete(AdminRequiredMixin, LoginRequiredMixin, DeleteView):
    model = OfficeCategory
    template_name = 'office_supplies/category_confirm_delete.html'
    success_url = reverse_lazy('office_supplies:category-list')

class OfficeSupplyDetail(LoginRequiredMixin, DetailView):
    model = OfficeSupply
    template_name = 'office_supplies/detail.html'


class OfficeMassOutgoingCreateView(AdminRequiredMixin, LoginRequiredMixin, View):
    template_name = 'office_supplies/mass_outgoing_form.html'
    success_url = reverse_lazy('office_supplies:supply-list')

    def get(self, request):
        return render(request, self.template_name, {
            'form': OfficeMassOutgoingForm(),
            'formset': OfficeMassOutgoingItemFormSet(prefix='outgoing_items')
        })

    def post(self, request):
        form = OfficeMassOutgoingForm(request.POST)
        formset = OfficeMassOutgoingItemFormSet(request.POST, prefix='outgoing_items')

        if (form.is_valid() and formset.is_valid()):
            try:
                with transaction.atomic():
                    reason = form.cleaned_data.get('reason')

                    # Validate all items first
                    items_to_process = []
                    for item_form in formset:
                        if (item_form.is_valid() and item_form.cleaned_data and not item_form.cleaned_data.get('DELETE', False)):
                            supply = item_form.cleaned_data.get('supply')
                            quantity = item_form.cleaned_data.get('quantity')
                            
                            if (not supply or not quantity):
                                continue
                                
                            if (supply.quantity < quantity):
                                raise ValidationError(f"Insufficient stock for {supply.name}")
                            
                            items_to_process.append({
                                'supply': supply,
                                'quantity': quantity,
                                'previous_quantity': supply.quantity
                            })
                    
                    # Process all items after validation
                    for item in items_to_process:
                        supply = item['supply']
                        quantity = item['quantity']
                        previous_quantity = item['previous_quantity']
                        
                        # Update supply quantity
                        supply.quantity = previous_quantity - quantity
                        supply.save()
                        
                        # Create movement record
                        InventoryMovement.objects.create(
                            content_type=ContentType.objects.get_for_model(supply),
                            object_id=supply.id,
                            movement_type='OUT',
                            quantity=quantity,
                            processed_by=request.user,
                            notes=f"Reason: {reason}",  # Removed notes from here
                            previous_quantity=previous_quantity,
                            current_quantity=supply.quantity
                        )

                    messages.success(request, 'Outgoing transaction processed successfully')
                    return redirect(self.success_url)

            except ValidationError as e:
                messages.error(request, str(e))
            except Exception as e:
                messages.error(request, f'Error processing transaction: {str(e)}')

        return render(request, self.template_name, {
            'form': form,
            'formset': formset
        })

class ExpiredItemDeleteView(LoginRequiredMixin, DeleteView):
    model = OfficeSupply
    template_name = 'office_supplies/expired_confirm_delete.html'
    success_url = reverse_lazy('office_supplies:expired-list')

class ExpiredItemListView(LoginRequiredMixin, ListView):
    model = OfficeSupply
    template_name = 'office_supplies/expired_list.html'
    context_object_name = 'expired_items'

    def get_queryset(self):
        return OfficeSupply.objects.filter(expiration_date__lte=timezone.now().date())

class MassAddView(AdminRequiredMixin, LoginRequiredMixin, View):
    template_name = 'office_supplies/mass_add.html'
    
    def get(self, request):
        return render(request, self.template_name, {'form': MassAddForm()})

    @transaction.atomic
    def post(self, request):
        form = MassAddForm(request.POST, request.FILES)
        if (not form.is_valid()):
            return render(request, self.template_name, {'form': form})

        success_count = 0
        error_messages = []
        
        try:
            csv_file = request.FILES['csv_file']
            # Decode file content
            for encoding in ['utf-8-sig', 'utf-8', 'latin1']:
                try:
                    decoded_file = csv_file.read().decode(encoding)
                    csv_file.seek(0)
                    break
                except UnicodeDecodeError:
                    continue

            reader = csv.DictReader(io.StringIO(decoded_file))
            required_fields = ['name', 'quantity', 'category']
            
            # Validate CSV structure
            if (not all(field.lower() in [h.lower() for h in reader.fieldnames] for field in required_fields)):
                raise ValueError(f"Missing required fields: {', '.join(required_fields)}")

            # Process each row in a single transaction
            for row_num, row in enumerate(reader, start=2):
                try:
                    with transaction.atomic():
                        # Clean data
                        name = row.get('name', '').strip()
                        quantity = row.get('quantity', '').strip()
                        category = row.get('category', '').strip()
                        reorder_level = row.get('reorder_level', '10').strip()

                        # Validate required fields
                        if (not all([name, quantity, category])):
                            raise ValueError("Missing required data")

                        # Convert numeric values
                        try:
                            quantity = int(quantity)
                            reorder_level = int(reorder_level) if reorder_level else 10
                        except ValueError:
                            raise ValueError("Invalid number format")

                        if (quantity < 0 or reorder_level < 0):
                            raise ValueError("Numbers cannot be negative")

                        # Get or create category
                        category_obj, _ = OfficeCategory.objects.get_or_create(
                            name=category,
                            defaults={'description': 'Added via mass import'}
                        )

                        # Create or update supply
                        supply, created = OfficeSupply.objects.get_or_create(
                            name=name,
                            defaults={
                                'category': category_obj,
                                'reorder_level': reorder_level,
                            }
                        )

                        # Update quantity and create movement record
                        prev_quantity = supply.quantity
                        supply.quantity += quantity
                        supply.save()

                        # Create movement record
                        InventoryMovement.objects.create(
                            content_type=ContentType.objects.get_for_model(supply),
                            object_id=supply.id,
                            movement_type='IN',
                            quantity=quantity,
                            processed_by=request.user,
                            previous_quantity=prev_quantity,
                            current_quantity=supply.quantity,
                            notes=f"Mass add import - Row {row_num}"
                        )

                        success_count += 1

                except Exception as e:
                    error_messages.append(f"Row {row_num}: {str(e)}")
                    continue

            if (success_count > 0):
                messages.success(request, f'Successfully added {success_count} items')
            if (error_messages):
                messages.warning(request, f'Encountered {len(error_messages)} errors: {"; ".join(error_messages)}')

            return redirect('office_supplies:supply-list')

        except Exception as e:
            messages.error(request, f'Error processing file: {str(e)}')
            return render(request, self.template_name, {'form': form})

class DownloadCSVTemplateView(View):
    def get(self, request):
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="mass_add_template.csv"'
        response.write(u'\ufeff')  # BOM for Excel

        writer = csv.writer(response)
        writer.writerow(['name', 'category', 'quantity', 'expiry date', 'reorder_level', 'description'])
        writer.writerow(['Paper Clips', 'Office Supplies', '100', '31/12/2025', '10', 'Standard clips'])
        
        return response

class ExpiringSoonListView(LoginRequiredMixin, ListView):
    model = OfficeSupply
    template_name = 'office_supplies/expiring_soon_list.html'
    context_object_name = 'supplies'

    def get_queryset(self):
        today = timezone.now().date()
        return OfficeSupply.objects.select_related('category').filter(
            expiration_date__gt=today,
            expiration_date__lte=today + timedelta(days=30)
        ).order_by('expiration_date')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['now'] = timezone.now().date()
        return context

from django.http import HttpResponse
from reportlab.pdfgen import canvas
from django.utils import timezone

class OfficeReportView(View):
    report_types = {
        'expired': lambda: OfficeSupply.objects.filter(expiration_date__lte=timezone.now().date()),
        'low_stock': lambda: OfficeSupply.objects.filter(quantity__lte=F('reorder_level')),
        'all': OfficeSupply.objects.all
    }

    def get(self, request, *args, **kwargs):
        report_type = request.GET.get('type', 'all')
        format = request.GET.get('format', 'pdf')
        
        queryset = self.report_types.get(report_type, self.report_types['all'])()
        
        if (format == 'pdf'):
            return self.generate_pdf(queryset, report_type)
        elif (format == 'excel'):
            return self.generate_excel(queryset, report_type)
        return HttpResponse("Invalid format", status=400)

    def generate_excel(self, queryset, report_type):
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"office_report_{report_type}_{timezone.now().date()}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        wb = Workbook()
        ws = wb.active
        ws.title = f"{report_type.title()} Report"

        # Headers
        headers = ['Name', 'Category', 'Quantity', 'Reorder Level', 'Status']
        ws.append(headers)

        # Data
        for item in queryset:
            ws.append([
                item.name,
                item.category.name,
                item.quantity,
                item.reorder_level,
                'Low Stock' if item.needs_reorder else 'OK'
            ])

        wb.save(response)
        return response

    def generate_pdf(self, queryset, report_type):
        # Create the response object for the PDF
        response = HttpResponse(content_type='application/pdf')
        filename = f"office_report_{report_type}_{timezone.now().date()}.pdf"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        # Create a canvas object using the response
        p = canvas.Canvas(response)
        
        # Add some header text to the PDF
        p.setFont("Helvetica-Bold", 16)
        p.drawString(50, 800, f"Office Supplies Report - {report_type.replace('_', ' ').title()}")

        # Set the font for the rest of the text
        p.setFont("Helvetica", 12)

        y = 750
        for item in queryset:
            p.drawString(50, y, f"Name: {item.name}")
            p.drawString(200, y, f"Category: {item.category.name}")
            p.drawString(350, y, f"Quantity: {item.quantity}")
            p.drawString(500, y, f"Reorder Level: {item.reorder_level}")
            p.drawString(650, y, f"Status: {'Low Stock' if item.needs_reorder else 'OK'}")
            y -= 20

            # If we are near the bottom of the page, create a new page
            if (y < 50):
                p.showPage()
                y = 800

        p.showPage()
        p.save()

        return response

class LowStockListView(LoginRequiredMixin, ListView):
    model = OfficeSupply
    template_name = 'office_supplies/low_stock_list.html'
    context_object_name = 'supplies'

    def get_queryset(self):
        return OfficeSupply.objects.select_related('category').filter(
            quantity__lte=F('reorder_level')
        ).order_by('quantity')

class OfficeReportsView(LoginRequiredMixin, TemplateView):
    template_name = 'office_supplies/reports.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Get date range for filtering
        end_date = timezone.now()
        start_date = end_date - timedelta(days=30)

        # Fetch office-specific data
        supplies = OfficeSupply.objects.select_related('category')
        
        # Add office-specific metrics
        context.update({
            'title': 'Office Supplies Inventory Report',
            'total_supplies': supplies.count(),
            'low_stock_count': supplies.filter(quantity__lte=F('reorder_level')).count(),
            'supplies': supplies.order_by('-last_updated')[:10],  # Most recently updated
            'low_stock': supplies.filter(quantity__lte=F('reorder_level'))[:5],
            'categories_summary': self.get_categories_summary(supplies),
            'monthly_usage': self.get_monthly_usage(),
            'start_date': start_date,
            'end_date': end_date,
        })
        
        return context

    def get_categories_summary(self, supplies):
        return supplies.values('category__name').annotate(
            total_items=Count('id'),
            total_quantity=Sum('quantity'),
            low_stock_items=Count(
                'id',
                filter=Q(quantity__lte=F('reorder_level'))
            )
        )

    def get_monthly_usage(self):
        thirty_days_ago = timezone.now() - timedelta(days=30)
        return OfficeSupply.objects.filter(
            last_updated__gte=thirty_days_ago
        ).values('name').annotate(
            usage=F('initial_quantity') - F('quantity')
        ).order_by('-usage')[:5]

class CategoryListView(LoginRequiredMixin, ListView):
    model = OfficeCategory
    template_name = 'office_supplies/category_list.html'
    context_object_name = 'categories'

    def get_queryset(self):
        queryset = super().get_queryset().annotate(
            item_count=Count('officesupply'),
            low_stock_count=Count(
                'officesupply',
                filter=Q(officesupply__quantity__lte=F('officesupply__reorder_level'))
            )
        )
        search_query = self.request.GET.get('q')
        if (search_query):
            queryset = queryset.filter(name__icontains=search_query)
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update({
            'title': 'Office Supply Categories',
            'total_items': OfficeSupply.objects.count(),
            'low_stock_count': OfficeSupply.objects.filter(
                quantity__lte=F('reorder_level')
            ).count()
        })
        return context

from django.contrib import admin
from django.urls import path, include
from django.views.generic import RedirectView

urlpatterns = [
    path('admin/', admin.site.urls),
    path('', RedirectView.as_view(url='/vet/', permanent=False)),
    path('vet/', include('vet_supplies.urls')),
    path('office/', include('office_supplies.urls')),
    path('login/', LoginView.as_view(), name='login'),
    path('logout/', LogoutView.as_view(), name='logout'),
]

@login_required
def bulk_import_view(request):
    if request.method == 'POST':
        try:
            excel_file = request.FILES['excel_file']
            wb = load_workbook(excel_file, data_only=True)  # Added data_only=True to handle formulas
            ws = wb.active
            success_count = 0
            error_messages = []
            
            with transaction.atomic():
                for row_num, row in enumerate(ws.iter_rows(min_row=2), 2):
                    try:
                        # Get raw values and convert to proper types
                        name = str(row[0].value).strip() if row[0].value else None
                        category_name = str(row[1].value).strip() if row[1].value else None
                        quantity = int(float(row[2].value)) if row[2].value else 0
                        reorder_level = int(float(row[3].value)) if row[3].value else 10
                        expiry_date = row[4].value
                        description = str(row[5].value).strip() if row[5].value else ""

                        # Validate required fields
                        if not name or not category_name:
                            raise ValueError("Name and category are required")
                        
                        if quantity < 0:
                            raise ValueError("Quantity cannot be negative")

                        # Handle expiry date conversion with multiple format attempts
                        if expiry_date:
                            try:
                                # If it's already a datetime object
                                if isinstance(expiry_date, datetime):
                                    expiry_date = expiry_date.date()
                                # If it's already a date object
                                elif isinstance(expiry_date, date):
                                    expiry_date = expiry_date
                                # Handle Excel serial date
                                elif isinstance(expiry_date, (int, float)):
                                    # Convert Excel serial date to Python date
                                    # Excel serial dates start from 1900-01-01
                                    expiry_date = datetime(1899, 12, 30) + timedelta(days=int(expiry_date))
                                    expiry_date = expiry_date.date()
                                # If it's a string, try multiple formats    
                                elif isinstance(expiry_date, str):
                                    expiry_date = expiry_date.strip()
                                    for fmt in ['%d/%m/%Y', '%d-%m-%Y', '%Y-%m-%d']:
                                        try:
                                            expiry_date = datetime.strptime(expiry_date, fmt).date()
                                            break
                                        except ValueError:
                                            continue
                                    else:
                                        raise ValueError(f"Invalid date format for expiry date: {expiry_date}. Use DD/MM/YYYY format")
                                else:
                                    raise ValueError(f"Invalid date value: {expiry_date}")
                            except Exception as e:
                                raise ValueError(f"Date error: {str(e)}")
                        else:
                            expiry_date = None

                        # Create or get category
                        category, _ = OfficeCategory.objects.get_or_create(name=category_name)
                        
                        # Create or update supply
                        supply, created = OfficeSupply.objects.get_or_create(
                            name=name,
                            defaults={
                                'category': category,
                                'quantity': quantity,
                                'reorder_level': reorder_level,
                                'expiration_date': expiry_date,
                                'description': description
                            }
                        )
                        
                        if not created:
                            supply.quantity = quantity
                            supply.reorder_level = reorder_level
                            supply.expiration_date = expiry_date
                            supply.description = description
                            supply.save()
                            
                        success_count += 1

                    except Exception as e:
                        error_messages.append(f"Row {row_num}: {str(e)}")
                        continue

            if success_count > 0:
                messages.success(request, f'Successfully imported {success_count} items')
            if error_messages:
                messages.warning(request, f'Encountered {len(error_messages)} errors: {"; ".join(error_messages)}')

            return redirect('office_supplies:supply-list')

        except Exception as e:
            messages.error(request, f'Error processing file: {str(e)}')

    return render(request, 'office_supplies/bulk_import.html')

@login_required
def download_template(request):
    format = request.GET.get('format', 'xlsx')
    
    if (format == 'csv'):
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="office_supplies_template.csv"'
        response.write(u'\ufeff')  # BOM for Excel

        writer = csv.writer(response)
        writer.writerow(['name', 'category', 'quantity', 'reorder_level', 'expiry_date', 'description'])
        # Add sample data
        writer.writerow(['Paper Clips', 'Office Supplies', 100, 10, '31/12/2025', 'Standard clips'])
        writer.writerow(['Staples', 'Stationery', 200, 50, '30/06/2025', 'Premium quality'])
        return response

    # Excel format
    wb = Workbook()
    ws = wb.active
    ws.title = "Import Template"
    
    # Add headers
    headers = ['name', 'category', 'quantity', 'reorder_level', 'expiry_date', 'description']
    ws.append(headers)
    
    # Add sample data
    sample_data = [
        ['Paper Clips', 'Office Supplies', 100, 10, '31/12/2025', 'Standard clips'],
        ['Staples', 'Stationery', 200, 50, '30/06/2025', 'Premium quality']
    ]
    for row in sample_data:
        ws.append(row)

    # Set column widths
    column_widths = [20, 15, 10, 12, 15, 40]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=office_supplies_template.xlsx'
    wb.save(response)
    return response

from django.contrib.auth.mixins import LoginRequiredMixin
from django.views.generic import ListView
from django.db.models import Q
from django.contrib.contenttypes.models import ContentType
from reports.models import InventoryMovement
from .models import OfficeSupply
from vet_supplies.models import VetSupply

class TransactionHistoryView(LoginRequiredMixin, ListView):
    template_name = 'office_supplies/transaction_history.html'
    context_object_name = 'transactions'
    paginate_by = 10

    def get_queryset(self):
        queryset = InventoryMovement.objects.select_related(
            'content_type', 
            'processed_by'
        ).prefetch_related('content_object')
        
        filters = {}
        
        if (self.request.GET.get('item_name')):
            vet_ct = ContentType.objects.get_for_model(VetSupply)
            office_ct = ContentType.objects.get_for_model(OfficeSupply)
            
            vet_ids = VetSupply.objects.filter(
                name__icontains=self.request.GET['item_name']
            ).values_list('id', flat=True)
            office_ids = OfficeSupply.objects.filter(
                name__icontains=self.request.GET['item_name']
            ).values_list('id', flat=True)
            
            queryset = queryset.filter(
                (Q(content_type=vet_ct) & Q(object_id__in=vet_ids)) |
                (Q(content_type=office_ct) & Q(object_id__in=office_ids))
            )

        # Filter by supply type
        supply_type = self.request.GET.get('supply_type')
        if (supply_type):
            if (supply_type == 'vet'):
                ct = ContentType.objects.get_for_model(VetSupply)
            elif (supply_type == 'office'):
                ct = ContentType.objects.get_for_model(OfficeSupply)
            if (supply_type in ['vet', 'office']):
                queryset = queryset.filter(content_type=ct)
        
        # Filter by movement type
        movement_type = self.request.GET.get('movement_type')
        if (movement_type):
            filters['movement_type'] = movement_type
            
        if (self.request.GET.get('date_from')):
            filters['timestamp__date__gte'] = self.request.GET['date_from']
            
        if (self.request.GET.get('date_to')):
            filters['timestamp__date__lte'] = self.request.GET['date_to']
        
        return queryset.filter(**filters).order_by('-timestamp')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['from_type'] = self.request.GET.get('from_type', '')
        
        # Add supply types
        context['supply_types'] = [
            ('vet', 'Veterinary'),
            ('office', 'Office')
        ]
        context['selected_supply_type'] = self.request.GET.get('supply_type', '')
        
        # Add movement types
        context['movement_types'] = InventoryMovement.MOVEMENT_TYPES
        context['selected_type'] = self.request.GET.get('movement_type', '')
        
        return context

@login_required
def export_transactions(request):
    view = TransactionHistoryView()
    view.request = request
    queryset = view.get_queryset()
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="transactions.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['Date', 'Supply Type', 'Item', 'Type', 'Quantity', 'Previous Qty', 'New Qty', 'Processed By', 'Notes'])
    
    for t in queryset:
        item = t.content_object
        supply_type = 'Veterinary' if isinstance(item, VetSupply) else 'Office'
        item_name = item.name if item else 'Unknown Item'
        
        writer.writerow([
            t.timestamp.strftime('%Y-%m-%d %H:%M'),
            supply_type,
            item_name,
            dict(InventoryMovement.MOVEMENT_TYPES)[t.movement_type],
            t.quantity,
            t.previous_quantity,
            t.current_quantity,
            t.processed_by.username if t.processed_by else 'System',
            t.notes
        ])
    
    return response

@login_required
def adjust_inventory(request):
    if request.method == 'POST':
        supply_id = request.POST.get('supply_id')
        adjustment = int(request.POST.get('adjustment', 0))
        reason = request.POST.get('reason', '')
        
        try:
            supply = OfficeSupply.objects.get(id=supply_id)
            previous_quantity = supply.quantity
            new_quantity = previous_quantity + adjustment
            
            if (new_quantity < 0):
                messages.error(request, "Adjustment would result in negative stock")
                return redirect('office_supplies:supply-list')
            
            supply.quantity = new_quantity
            supply.save()
            
            InventoryMovement.objects.create(
                content_type=ContentType.objects.get_for_model(supply),
                object_id=supply.id,
                movement_type='ADJ',
                quantity=abs(adjustment),
                processed_by=request.user,
                previous_quantity=previous_quantity,
                current_quantity=new_quantity,
                notes=f"Manual adjustment: {reason}. Change: {'+'if adjustment > 0 else ''}{adjustment}"
            )
            
            messages.success(request, f'Successfully adjusted {supply.name} quantity')
            
        except OfficeSupply.DoesNotExist:
            messages.error(request, "Supply not found")
        except Exception as e:
            messages.error(request, f"Error adjusting inventory: {str(e)}")
            
    return redirect('office_supplies:supply-list')

@login_required
def download_template(request):
    format = request.GET.get('format', 'xlsx')
    
    if format == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="office_supplies_template.csv"'
        response.write(u'\ufeff')  # BOM for Excel

        writer = csv.writer(response)
        writer.writerow(['name', 'category', 'quantity', 'reorder_level', 'expiry_date', 'description'])
        # Add sample data
        writer.writerow(['Paper Clips', 'Office Supplies', 100, 10, '31/12/2025', 'Standard clips'])
        writer.writerow(['Staples', 'Stationery', 200, 50, '30/06/2025', 'Premium quality'])
        return response

    # Excel format
    wb = Workbook()
    ws = wb.active
    ws.title = "Import Template"
    
    # Add headers
    headers = ['name', 'category', 'quantity', 'reorder_level', 'expiry_date', 'description']
    ws.append(headers)
    
    # Add sample data
    sample_data = [
        ['Paper Clips', 'Office Supplies', 100, 10, '31/12/2025', 'Standard clips'],
        ['Staples', 'Stationery', 200, 50, '30/06/2025', 'Premium quality']
    ]
    for row in sample_data:
        ws.append(row)

    # Set column widths
    column_widths = [20, 15, 10, 12, 15, 40]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=office_supplies_template.xlsx'
    wb.save(response)
    return response

@login_required
def bulk_import_view(request):
    if request.method == 'POST':
        try:
            excel_file = request.FILES['excel_file']
            wb = load_workbook(excel_file, data_only=True)  # Added data_only=True to handle formulas
            ws = wb.active
            success_count = 0
            error_messages = []
            
            with transaction.atomic():
                for row_num, row in enumerate(ws.iter_rows(min_row=2), 2):
                    try:
                        # Get raw values and convert to proper types
                        name = str(row[0].value).strip() if row[0].value else None
                        category_name = str(row[1].value).strip() if row[1].value else None
                        quantity = int(float(row[2].value)) if row[2].value else 0
                        reorder_level = int(float(row[3].value)) if row[3].value else 10
                        expiry_date = row[4].value
                        description = str(row[5].value).strip() if row[5].value else ""

                        # Validate required fields
                        if not name or not category_name:
                            raise ValueError("Name and category are required")
                        
                        if quantity < 0:
                            raise ValueError("Quantity cannot be negative")

                        # Handle expiry date conversion with multiple format attempts
                        if expiry_date:
                            try:
                                # If it's already a datetime object
                                if isinstance(expiry_date, datetime):
                                    expiry_date = expiry_date.date()
                                # If it's already a date object
                                elif isinstance(expiry_date, date):
                                    expiry_date = expiry_date
                                # Handle Excel serial date
                                elif isinstance(expiry_date, (int, float)):
                                    # Convert Excel serial date to Python date
                                    # Excel serial dates start from 1900-01-01
                                    expiry_date = datetime(1899, 12, 30) + timedelta(days=int(expiry_date))
                                    expiry_date = expiry_date.date()
                                # If it's a string, try multiple formats    
                                elif isinstance(expiry_date, str):
                                    expiry_date = expiry_date.strip()
                                    for fmt in ['%d/%m/%Y', '%d-%m-%Y', '%Y-%m-%d']:
                                        try:
                                            expiry_date = datetime.strptime(expiry_date, fmt).date()
                                            break
                                        except ValueError:
                                            continue
                                    else:
                                        raise ValueError(f"Invalid date format for expiry date: {expiry_date}. Use DD/MM/YYYY format")
                                else:
                                    raise ValueError(f"Invalid date value: {expiry_date}")
                            except Exception as e:
                                raise ValueError(f"Date error: {str(e)}")
                        else:
                            expiry_date = None

                        # Create or get category
                        category, _ = OfficeCategory.objects.get_or_create(name=category_name)
                        
                        # Create or update supply
                        supply, created = OfficeSupply.objects.get_or_create(
                            name=name,
                            defaults={
                                'category': category,
                                'quantity': quantity,
                                'reorder_level': reorder_level,
                                'expiration_date': expiry_date,
                                'description': description
                            }
                        )
                        
                        if not created:
                            supply.quantity = quantity
                            supply.reorder_level = reorder_level
                            supply.expiration_date = expiry_date
                            supply.description = description
                            supply.save()
                            
                        success_count += 1

                    except Exception as e:
                        error_messages.append(f"Row {row_num}: {str(e)}")
                        continue

            if success_count > 0:
                messages.success(request, f'Successfully imported {success_count} items')
            if error_messages:
                messages.warning(request, f'Encountered {len(error_messages)} errors: {"; ".join(error_messages)}')

            return redirect('office_supplies:supply-list')

        except Exception as e:
            messages.error(request, f'Error processing file: {str(e)}')

    return render(request, 'office_supplies/bulk_import.html')