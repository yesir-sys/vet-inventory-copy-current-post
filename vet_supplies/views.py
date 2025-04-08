from django import forms
from django.forms import ValidationError
from reportlab.pdfgen import canvas  # Correct import for PDF generation
from django.views import View
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import csv
from io import TextIOWrapper
from django.utils import timezone
from django.urls import reverse_lazy
from django.template.loader import render_to_string
from django.views.generic import ListView, CreateView, UpdateView, DeleteView, DetailView, TemplateView
from django.contrib.auth.views import LoginView, LogoutView
from django.contrib.auth.forms import UserCreationForm
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth import login
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.db import IntegrityError, transaction
from django.db.models import Q, F, Sum
from django.http import HttpResponse, JsonResponse, HttpResponseRedirect
from .models import VetSupply, VetCategory, ExpiredItem, OfficeSupply
from .forms import MassOutgoingItemFormSet, VetSupplyForm, MassOutgoingForm, MassAddForm
from datetime import date, timedelta
from django.db.models.deletion import ProtectedError
from django.db import connection
from django.contrib.auth.decorators import login_required
import io
from datetime import datetime
from django.db.models import Q
from django.views.generic import ListView
from reports.models import InventoryMovement  # Add this import at the top
from django.contrib.contenttypes.models import ContentType  # Add this import at the top
from core.mixins import AdminRequiredMixin


# Authentication Views
class UserSignUpView(CreateView):
    form_class = UserCreationForm
    template_name = 'registration/signup.html'
    success_url = reverse_lazy('vet_supplies:supply-list')

    def form_valid(self, form):
        response = super().form_valid(form)
        login(self.request, self.object)
        return response

class UserLoginView(LoginView):
    template_name = 'registration/login.html'

class UserLogoutView(LogoutView):
    next_page = reverse_lazy('vet_supplies:supply-list')

# Vet Supply Views
class VetSupplyList(LoginRequiredMixin, ListView):
    model = VetSupply
    template_name = 'vet_supplies/list.html'
    context_object_name = 'supplies'
    paginate_by = 10

    def get_queryset(self):
        queryset = VetSupply.objects.select_related('category').all()
        search_query = self.request.GET.get('q')
        category_id = self.request.GET.get('category')
        stock_status = self.request.GET.get('stock_status')
        expiration_status = self.request.GET.get('expiration_status')

        if search_query:
            queryset = queryset.filter(
                Q(name__icontains=search_query) |
                Q(category__name__icontains=search_query)
            )
        
        if category_id and category_id.isdigit():
            queryset = queryset.filter(category_id=category_id)
            
        if stock_status == 'low':
            queryset = queryset.filter(quantity__lte=F('reorder_level'))
        elif stock_status == 'normal':
            queryset = queryset.filter(quantity__gt=F('reorder_level'))

        # Handle expiration status filtering
        today = timezone.now().date()
        if expiration_status:
            if expiration_status == 'expired':
                queryset = queryset.filter(expiration_date__lt=today)
            elif expiration_status == 'expiring_soon':
                thirty_days_later = today + timedelta(days=30)
                queryset = queryset.filter(
                    expiration_date__gt=today,
                    expiration_date__lte=thirty_days_later
                )
            elif expiration_status == 'not_expiring':
                thirty_days_later = today + timedelta(days=30)
                queryset = queryset.filter(
                    Q(expiration_date__gt=thirty_days_later) | 
                    Q(expiration_date__isnull=True)
                )

        return queryset.distinct()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        today = timezone.now().date()
        thirty_days_later = today + timedelta(days=30)
        
        context.update({
            'categories': VetCategory.objects.all(),
            'selected_category': self.request.GET.get('category', ''),
            'low_stock_count': VetSupply.objects.filter(
                quantity__lte=F('reorder_level')
            ).count(),
                'expiring_soon_count': VetSupply.objects.filter(
                expiration_date__gt=today,
                expiration_date__lte=thirty_days_later
            ).count(),
            'search_query': self.request.GET.get('q', ''),
            'selected_category': int(self.request.GET.get('category', 0) or 0),
            'stock_status': self.request.GET.get('stock_status', ''),
            'expiration_status': self.request.GET.get('expiration_status', '')
        })
        return context

# API endpoint to get supplies as JSON
def get_supplies_json(request):
    supplies = VetSupply.objects.all().values('name', 'quantity', 'reorder_level', 'expiration_date', 'expiration_status')
    return JsonResponse(list(supplies), safe=False)


class VetSupplyCreate(AdminRequiredMixin, LoginRequiredMixin, CreateView):
    model = VetSupply
    form_class = VetSupplyForm
    template_name = 'vet_supplies/form.html'

    def get_success_url(self):
        return reverse_lazy('vet_supplies:supply-list')

class VetSupplyUpdate(AdminRequiredMixin, LoginRequiredMixin, UpdateView):
    model = VetSupply
    form_class = VetSupplyForm
    template_name = 'vet_supplies/form.html'

    def get_success_url(self):
        return reverse_lazy('vet_supplies:supply-list')

    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(self.request, f'{self.object.name} updated successfully')
        return response

class VetSupplyDelete(AdminRequiredMixin, LoginRequiredMixin, DeleteView):
    model = VetSupply
    template_name = 'vet_supplies/confirm_delete.html'
    success_url = reverse_lazy('vet_supplies:supply-list')

    def delete(self, request, *args, **kwargs):
        self.object = self.get_object()
        try:
            self.object.delete()
            messages.success(request, 'Supply deleted successfully')
            return redirect(self.success_url)
        except ProtectedError:
            messages.error(request, 'Cannot delete this supply because it is referenced by other records')
            return redirect(self.success_url)

# Category Views
class VetCategoryList(LoginRequiredMixin, ListView):
    model = VetCategory
    template_name = 'vet_supplies/category_list.html'

class VetCategoryCreate(AdminRequiredMixin, LoginRequiredMixin, CreateView):
    model = VetCategory
    template_name = 'vet_supplies/category_form.html'
    fields = ['name', 'description']
    success_url = reverse_lazy('vet_supplies:category-list')

class VetCategoryUpdate(AdminRequiredMixin, LoginRequiredMixin, UpdateView):
    model = VetCategory
    fields = ['name']
    template_name = 'vet_supplies/category_form.html'
    success_url = reverse_lazy('vet_supplies:category-list')

class VetCategoryDelete(AdminRequiredMixin, LoginRequiredMixin, DeleteView):
    model = VetCategory
    success_url = reverse_lazy('vet_supplies:category-list')
    template_name = 'vet_supplies/vetcategory_confirm_delete.html'

    def delete(self, request, *args, **kwargs):
        category = self.get_object()
        messages.success(request, f'Category "{category.name}" was successfully deleted.')
        return super().delete(request, *args, **kwargs)

class VetSupplyDetail(LoginRequiredMixin, DetailView):
    model = VetSupply
    template_name = 'vet_supplies/detail.html'

class ExpiredItemDeleteView(LoginRequiredMixin, DeleteView):
    model = ExpiredItem
    template_name = 'vet_supplies/confirm_delete.html'
    success_url = reverse_lazy('vet_supplies:expired-list')

    def delete(self, request, *args, **kwargs):
        messages.success(request, 'Expired item removed successfully')
        return super().delete(request, *args, **kwargs)

# Expired Item List View
class ExpiredItemListView(LoginRequiredMixin, ListView):
    model = ExpiredItem
    template_name = 'vet_supplies/expired_list.html'
    context_object_name = 'items'

    def get_queryset(self):
        return ExpiredItem.objects.all()

class MassAddView(AdminRequiredMixin, LoginRequiredMixin, View):
    template_name = 'vet_supplies/mass_add.html'
    form_class = MassAddForm

    def get(self, request):
        return render(request, self.template_name, {'form': self.form_class()})

    @transaction.atomic
    def post(self, request):
        form = self.form_class(request.POST, request.FILES)
        if not form.is_valid():
            return render(request, self.template_name, {'form': form})

        success_count = 0
        error_messages = []
        
        try:
            csv_file = request.FILES['csv_file']
            decoded_file = None
            
            # Handle different file encodings
            for encoding in ['utf-8-sig', 'utf-8', 'latin1']:
                try:
                    decoded_file = csv_file.read().decode(encoding)
                    csv_file.seek(0)
                    break
                except UnicodeDecodeError:
                    continue
            
            if not decoded_file:
                raise ValueError("Could not decode file with supported encodings")

            reader = csv.DictReader(io.StringIO(decoded_file))
            
            # Validate header fields
            required_fields = ['name', 'quantity', 'category']
            missing_fields = [field for field in required_fields if field not in reader.fieldnames]
            if missing_fields:
                raise ValueError(f"Missing required fields: {', '.join(missing_fields)}")

            for row_num, row in enumerate(reader, start=2):
                try:
                    # Clean and validate data
                    name = row.get('name', '').strip()
                    category = row.get('category', '').strip()
                    
                    # Convert quantity to int, handle None/empty values
                    try:
                        quantity = int(row.get('quantity', 0) or 0)
                        reorder_level = int(row.get('reorder_level', 10) or 10)
                    except (ValueError, TypeError):
                        raise ValueError("Invalid number format for quantity or reorder level")

                    if not name or not category:
                        raise ValueError("Name and category are required")

                    # Convert quantity to 0 if None or negative
                    quantity = max(0, quantity)
                    reorder_level = max(0, reorder_level)
                    
                    # Get or create category
                    category_obj, _ = VetCategory.objects.get_or_create(name=category)
                    
                    # Create or update supply
                    supply, created = VetSupply.objects.get_or_create(
                        name=name,
                        defaults={
                            'category': category_obj,
                            'reorder_level': reorder_level,
                            'quantity': 0  # Start with 0
                        }
                    )
                    
                    # Update quantity and create movement record
                    previous_quantity = supply.quantity or 0  # Handle None
                    supply.quantity = previous_quantity + quantity
                    supply.save()
                    
                    # Create movement record
                    if quantity > 0:  # Only create movement record if quantity is positive
                        InventoryMovement.objects.create(
                            content_type=ContentType.objects.get_for_model(supply),
                            object_id=supply.id,
                            movement_type='IN',
                            quantity=quantity,
                            processed_by=request.user,
                            previous_quantity=previous_quantity,
                            current_quantity=supply.quantity,
                            notes=f"Mass add import - Row {row_num}"
                        )
                    
                    success_count += 1
                    
                except Exception as e:
                    error_messages.append(f"Row {row_num}: {str(e)}")
                    continue
            
            if success_count > 0:
                messages.success(request, f'Successfully added {success_count} items')
            if error_messages:
                messages.warning(request, f'Encountered {len(error_messages)} errors: {"; ".join(error_messages)}')
            
            return redirect('vet_supplies:supply-list')
            
        except Exception as e:
            messages.error(request, f'Error processing file: {str(e)}')
            return render(request, self.template_name, {'form': form})

class DownloadCSVTemplateView(View):
    def get(self, request, *args, **kwargs):
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="template.csv"'

        writer = csv.writer(response)
        writer.writerow(['Column1', 'Column2', 'Column3'])  # Add your template columns here

        return response

class ExpiringSoonListView(LoginRequiredMixin, ListView):
    model = VetSupply
    template_name = 'vet_supplies/expiring_soon.html'
    context_object_name = 'supplies'

    def get_queryset(self):
        thirty_days_from_now = timezone.now().date() + timedelta(days=30)
        return VetSupply.objects.filter(
            expiration_date__lte=thirty_days_from_now,
            expiration_date__gt=timezone.now().date()
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Items Expiring Soon'
        return context
    
    
# Low Stock List View to show items with low stock levels
class LowStockListView(LoginRequiredMixin, ListView):
    model = VetSupply
    template_name = 'vet_supplies/low_stock.html'
    context_object_name = 'supplies'

    def get_queryset(self):
        return VetSupply.objects.filter(quantity__lte=F('reorder_level'))

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Low Stock Items'
        return context

class ReportsView(LoginRequiredMixin, TemplateView):
    template_name = 'vet_supplies/reports.html'

    def get(self, request, *args, **kwargs):
        report_format = request.GET.get('format')
        report_type = request.GET.get('type')

        if report_format in ['pdf', 'excel']:
            queryset = VetSupply.objects.all()
            if report_type == 'low_stock':
                queryset = queryset.filter(quantity__lte=F('reorder_level'))
            elif report_type == 'expiring':
                queryset = queryset.filter(
                    expiration_date__lte=timezone.now().date() + timedelta(days=30),
                    expiration_date__gt=timezone.now().date()
                )
            
            if report_format == 'pdf':
                return self.generate_pdf(queryset, report_type)
            return self.generate_excel(queryset, report_type)
        
        return super().get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Get date range for filtering
        end_date = timezone.now()
        start_date = end_date - timedelta(days=30)

        # Fetch data with optimized queries using correct related name
        supplies = VetSupply.objects.select_related('category')\
            .annotate(total_used=Sum('vet_supply_transactions__quantity'))
        
        low_stock = supplies.filter(quantity__lte=F('reorder_level'))
        expiring_soon = supplies.filter(
            expiration_date__lte=timezone.now().date() + timedelta(days=30),
            expiration_date__gt=timezone.now().date()
        )

        context.update({
            'title': 'Inventory Reports',
            'total_supplies': supplies.count(),
            'low_stock_count': low_stock.count(),
            'expiring_soon_count': expiring_soon.count(),
            'supplies': supplies[:10],
            'low_stock': low_stock[:5],
            'expiring_soon': expiring_soon[:5],
            'start_date': start_date,
            'end_date': end_date,
        })
        
        return context

    def generate_pdf(self, queryset, report_type):
        response = HttpResponse(content_type='application/pdf')
        filename = f"inventory_report_{report_type}_{timezone.now().date()}.pdf"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        p = canvas.Canvas(response)
        p.setFont("Helvetica-Bold", 16)
        p.drawString(50, 800, f"Inventory Report ({report_type.replace('_', ' ').title()})")
        
        y = 750
        p.setFont("Helvetica", 12)
        for supply in queryset:
            p.drawString(50, y, f"{supply.name}")
            p.drawString(200, y, f"Quantity: {supply.quantity}")
            p.drawString(350, y, f"Exp: {supply.expiration_date or 'N/A'}")
            y -= 20
            if y < 50:
                p.showPage()
                y = 800
        p.save()
        return response

    def generate_excel(self, queryset, report_type):
        response = HttpResponse(content_type='application/ms-excel')
        filename = f"vet_report_{report_type}_{timezone.now().date()}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        wb = Workbook()
        ws = wb.active
        ws.title = f"{report_type} Report"
        
        ws.append([
            'Name', 'Category', 'Quantity',
            'Reorder Level', 'Expiration Date', 'Status'
        ])
        
        for supply in queryset:
            ws.append([
                supply.name,
                supply.category.name,
                supply.quantity,
                supply.reorder_level,
                supply.expiration_date or 'N/A',
                supply.expiration_status
            ])
        
        wb.save(response)
        return response

# Remove or comment out the old reports_view function
# def reports_view(request):
#     return render(request, 'vet_supplies/reports.html')

class VetReportView(LoginRequiredMixin, View):
    report_types = {
        'expired': lambda: VetSupply.objects.expired(),
        'low_stock': lambda: VetSupply.objects.filter(quantity__lte=F('reorder_level')),
        'all': VetSupply.objects.all
    }

    def get(self, request, *args, **kwargs):
        report_type = request.GET.get('type', 'all')
        format = request.GET.get('format', 'pdf')

        queryset = self.report_types.get(report_type, self.report_types['all'])()

        if format == 'pdf':
            return self.generate_pdf(queryset, report_type)
        elif format == 'excel':
            return self.generate_excel(queryset, report_type)
        return HttpResponse("Invalid format", status=400)

    def generate_pdf(self, queryset, report_type):
        response = HttpResponse(content_type='application/pdf')
        filename = f"vet_report_{report_type}_{timezone.now().date()}.pdf"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        # Correct instantiation of Canvas
        p = canvas.Canvas(response)
        p.setFont("Helvetica-Bold", 16)
        p.drawString(50, 800, f"Veterinary Supplies Report ({report_type.replace('_', ' ').title()})")

        y = 750
        p.setFont("Helvetica", 12)
        for supply in queryset:
            p.drawString(50, y, f"{supply.name}")
            p.drawString(200, y, f"Quantity: {supply.quantity}")
            p.drawString(350, y, f"Exp: {supply.expiration_date or 'N/A'}")
            y -= 20
            if y < 50:
                p.showPage()
                y = 800
        p.save()
        return response

    def generate_excel(self, queryset, report_type):
        response = HttpResponse(content_type='application/ms-excel')
        filename = f"vet_report_{report_type}_{timezone.now().date()}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        wb = Workbook()
        ws = wb.active
        ws.title = f"{report_type} Report"

        # Headers
        ws.append([
            'Name', 'Category', 'Quantity',
            'Reorder Level', 'Expiration Date', 'Status'
        ])

        # Data
        for supply in queryset:
            ws.append([
                supply.name,
                supply.category.name,
                supply.quantity,
                supply.reorder_level,
                supply.expiration_date or 'N/A',
                supply.expiration_status
            ])

        wb.save(response)
        return response

from django.http import JsonResponse
from django.urls import reverse_lazy
from django.views.generic.edit import DeleteView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib import messages
from django.db.models import ProtectedError

class CategoryDelete(LoginRequiredMixin, DeleteView):
    model = VetCategory
    success_url = reverse_lazy('vet_supplies:category-list')
    template_name = 'vet_supplies/vetcategory_confirm_delete.html'

    def delete(self, request, *args, **kwargs):
        category = self.get_object()
        messages.success(request, f'Category "{category.name}" was successfully deleted.')
        return super().delete(request, *args, **kwargs)

def home(request):
    return render(request, 'home.html')

@login_required
def mass_add(request):
    if request.method == 'POST':
        form = MassAddForm(request.POST, request.FILES)
        if form.is_valid():
            success_count = 0
            error_messages = []
            
            try:
                csv_file = request.FILES['csv_file']
                # Handle different file encodings
                for encoding in ['utf-8-sig', 'utf-8', 'latin1']:
                    try:
                        decoded_file = csv_file.read().decode(encoding)
                        csv_file.seek(0)  # Reset file pointer
                        break
                    except UnicodeDecodeError:
                        continue
                
                reader = csv.DictReader(io.StringIO(decoded_file))
                required_fields = ['name', 'quantity', 'category']
                
                if not all(field in reader.fieldnames for field in required_fields):
                    raise ValueError(f"Missing required fields: {', '.join(required_fields)}")
                
                with transaction.atomic():
                    for row_num, row in enumerate(reader, start=2):
                        try:
                            # Clean and validate data
                            name = row.get('name', '').strip()
                            quantity = row.get('quantity', '').strip()
                            category = row.get('category', '').strip()
                            reorder_level = row.get('reorder_level', '10').strip()
                            exp_date = row.get('expiration_date', '').strip()
                            
                            if not all([name, quantity, category]):
                                raise ValueError("Missing required data")
                            
                            try:
                                quantity = int(quantity)
                                reorder_level = int(reorder_level) if reorder_level else 10
                            except ValueError:
                                raise ValueError("Invalid number format")

                            if quantity < 0 or reorder_level < 0:
                                raise ValueError("Numbers cannot be negative")
                            
                            # Handle expiration date
                            expiration_date = None
                            if exp_date:
                                try:
                                    # Support multiple date formats
                                    for date_format in ['%d-%m-%Y', '%Y-%m-%d', '%d/%m/%Y']:
                                        try:
                                            expiration_date = datetime.strptime(exp_date, date_format).date()
                                            break
                                        except ValueError:
                                            continue
                                    if not expiration_date:
                                        raise ValueError("Invalid date format")
                                except ValueError as e:
                                    raise ValueError(f"Invalid date format: {exp_date}")
                            
                            # Get or create category
                            category_obj, _ = VetCategory.objects.get_or_create(name=category)
                            
                            # Create or update supply
                            supply, created = VetSupply.objects.update_or_create(
                                name=name,
                                defaults={
                                    'category': category_obj,
                                    'reorder_level': reorder_level,
                                    'expiration_date': expiration_date
                                }
                            )
                            
                            # Update quantity and create movement record
                            previous_quantity = supply.quantity
                            supply.quantity += quantity
                            supply.save()
                            
                            # Create inventory movement record
                            InventoryMovement.objects.create(
                                content_type=ContentType.objects.get_for_model(supply),
                                object_id=supply.id,
                                movement_type='IN',
                                quantity=quantity,
                                processed_by=request.user,
                                previous_quantity=previous_quantity,
                                current_quantity=supply.quantity,
                                notes=f"Mass add import - Row {row_num}"
                            )
                            
                            success_count += 1
                        
                        except Exception as e:
                            error_messages.append(f"Row {row_num}: {str(e)}")
                    
                    if success_count > 0:
                        messages.success(request, f'Successfully added {success_count} items')
                    if error_messages:
                        messages.warning(request, f'Encountered {len(error_messages)} errors: {"; ".join(error_messages)}')
                    
                    return redirect('vet_supplies:supply-list')
                    
            except Exception as e:
                messages.error(request, f'Error processing file: {str(e)}')
    else:
        form = MassAddForm()
    
    return render(request, 'vet_supplies/mass_add.html', {'form': form})

@login_required 
def download_template(request):
    format = request.GET.get('format', 'xlsx')
    
    if format == 'csv':
        response = HttpResponse(
            content_type='text/csv',
            headers={'Content-Disposition': 'attachment; filename="vet_supplies_template.csv"'},
        )
        response.write(u'\ufeff')
        
        writer = csv.writer(response)
        writer.writerow(['name', 'quantity', 'category', 'reorder_level', 'expiration_date', 'dosage'])
        
        # Example data with future dates in DD/MM/YYYY format
        writer.writerow(['Amoxicillin', '100', 'Antibiotics', '50', '31/12/2025', '250mg'])
        writer.writerow(['Syringes', '200', 'Medical Supplies', '100', '30/06/2025', ''])
        writer.writerow(['Bandages', '300', 'First Aid', '150', '31/12/2025', ''])
        
        return response
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Import Template"
        
        headers = ['Name', 'Category', 'Quantity', 'Reorder Level', 'Expiry Date', 'Dosage']
        ws.append(headers)
        
        # Add sample data with future dates in DD/MM/YYYY format
        sample_data = [
            ['Amoxicillin', 'Antibiotics', '100', '50', '31/12/2025', '250mg'],
            ['Syringes', 'Medical Supplies', '200', '100', '30/06/2025', '']
        ]
        for row in sample_data:
            ws.append(row)
        
        # Set column widths
        column_widths = [20, 15, 10, 12, 15, 12]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=bulk_import_template.xlsx'
        
        wb.save(response)
        return response

class MassOutgoingCreateView(AdminRequiredMixin, LoginRequiredMixin, View):
    template_name = 'vet_supplies/mass_outgoing_form.html'
    success_url = reverse_lazy('vet_supplies:supply-list')

    def get(self, request):
        return render(request, self.template_name, {
            'form': MassOutgoingForm(),
            'formset': MassOutgoingItemFormSet(prefix='items')  # Add prefix here
        })

    def post(self, request):
        form = MassOutgoingForm(request.POST)
        formset = MassOutgoingItemFormSet(request.POST, prefix='items')  # Add prefix here

        if form.is_valid() and formset.is_valid():
            try:
                with transaction.atomic():
                    reason = form.cleaned_data.get('reason')
                    notes = form.cleaned_data.get('notes', '')

                    # Validate all items first
                    items_to_process = []
                    for item_form in formset:
                        if item_form.is_valid() and item_form.cleaned_data and not item_form.cleaned_data.get('DELETE', False):
                            supply = item_form.cleaned_data.get('supply')
                            quantity = item_form.cleaned_data.get('quantity')
                            
                            if not supply or not quantity:
                                continue
                                
                            if supply.quantity < quantity:
                                raise ValidationError(f"Insufficient stock for {supply.name}")
                            
                            items_to_process.append({
                                'supply': supply,
                                'quantity': quantity,
                                'previous_quantity': supply.quantity
                            })
                    
                    # Process all items after validation
                    for item in items_to_process:
                        supply = item['supply']
                        quantity = item['quantity']
                        previous_quantity = item['previous_quantity']
                        
                        # Update supply quantity
                        supply.quantity = previous_quantity - quantity
                        supply.save()
                        
                        # Create movement record
                        InventoryMovement.objects.create(
                            content_type=ContentType.objects.get_for_model(supply),
                            object_id=supply.id,
                            movement_type='OUT',
                            quantity=quantity,
                            processed_by=request.user,
                            notes=f"Reason: {reason}\nNotes: {notes}".strip(),
                            previous_quantity=previous_quantity,
                            current_quantity=supply.quantity
                        )

                    messages.success(request, 'Outgoing transaction processed successfully')
                    return redirect(self.success_url)

            except ValidationError as e:
                messages.error(request, str(e))
            except Exception as e:
                messages.error(request, f'Error processing transaction: {str(e)}')

        return render(request, self.template_name, {
            'form': form, 
            'formset': formset
        })

class SupplyItemForm(forms.Form):
    supply = forms.ModelChoiceField(
        queryset=VetSupply.objects.all(),
        empty_label="Select an item"
    )
    quantity = forms.IntegerField(
        min_value=1,
        widget=forms.NumberInput(attrs={'min': '1'})
    )

@login_required
def bulk_import_view(request):
    if request.method == 'POST':
        try:
            excel_file = request.FILES['excel_file']
            wb = load_workbook(excel_file)
            ws = wb.active
            success_count = 0
            error_messages = []
            
            with transaction.atomic():
                for row in ws.iter_rows(min_row=2):
                    try:
                        name = row[0].value
                        category_name = row[1].value
                        quantity = int(row[2].value or 0)
                        reorder_level = int(row[3].value or 10)
                        expiry_date = row[4].value  # This can be string or datetime
                        
                        # Handle expiry date conversion
                        if expiry_date:
                            if isinstance(expiry_date, datetime):
                                expiry_date = expiry_date.date()
                            elif isinstance(expiry_date, str):
                                try:
                                    # Try DD/MM/YYYY format first
                                    expiry_date = datetime.strptime(expiry_date.strip(), '%d/%m/%Y').date()
                                except ValueError:
                                    try:
                                        # Try DD-MM-YYYY format
                                        expiry_date = datetime.strptime(expiry_date.strip(), '%d-%m-%Y').date()
                                    except ValueError:
                                        try:
                                            # Try YYYY-MM-DD format
                                            expiry_date = datetime.strptime(expiry_date.strip(), '%Y-%m-%d').date()
                                        except ValueError:
                                            raise ValueError(f"Invalid date format for expiry date: {expiry_date}. Use DD/MM/YYYY format.")
                        else:
                            expiry_date = None

                        # Validate expiry date is in the future
                        if expiry_date and expiry_date <= timezone.now().date():
                            raise ValueError("Expiration date must be in the future")

                        # Rest of the code remains the same
                        if not all([name, category_name]):
                            continue
                            
                        category, _ = VetCategory.objects.get_or_create(name=category_name.strip())
                        
                        supply, created = VetSupply.objects.get_or_create(
                            name=name.strip(),
                            defaults={
                                'category': category,
                                'reorder_level': reorder_level,
                                'expiration_date': expiry_date
                            }
                        )
                        
                        # Update expiry date even if supply exists
                        if not created and expiry_date:
                            supply.expiration_date = expiry_date
                            
                        if quantity > 0:
                            prev_quantity = supply.quantity
                            supply.quantity += quantity
                            supply.save()
                            
                            InventoryMovement.objects.create(
                                content_type=ContentType.objects.get_for_model(supply),
                                object_id=supply.id,
                                movement_type='IN',
                                quantity=quantity,
                                processed_by=request.user,
                                previous_quantity=prev_quantity,
                                current_quantity=supply.quantity,
                                notes=f"Bulk import"
                            )
                            
                            success_count += 1
                            
                    except Exception as e:
                        error_messages.append(f"Row {row[0].row}: {str(e)}")
                        continue
                        
            if success_count > 0:
                messages.success(request, f'Successfully imported {success_count} items')
            if error_messages:
                messages.warning(request, f'Encountered {len(error_messages)} errors')
                
            return redirect('vet_supplies:supply-list')
            
        except Exception as e:
            messages.error(request, f'Error processing file: {str(e)}')
            
    return render(request, 'vet_supplies/bulk_import.html')

@login_required
def download_template(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Import Template"
    
    # Add headers
    headers = ['Name', 'Category', 'Quantity', 'Reorder Level', 'Expiry Date']  # Added Expiry Date
    ws.append(headers)
    
    # Add sample data with medical supply examples
    sample_data = [
        ['Amoxicillin', 'Antibiotics', '100', '50', '31/12/2025'],
        ['Syringes', 'Medical Supplies', '200', '100', '30/06/2026']
    ]
    for row in sample_data:
        ws.append(row)
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=bulk_import_template.xlsx'
    
    wb.save(response)
    return response

# ...existing code...

class TransactionHistoryView(LoginRequiredMixin, ListView):
    template_name = 'vet_supplies/transaction_history.html'
    context_object_name = 'transactions'
    paginate_by = 10

    def get_queryset(self):
        queryset = InventoryMovement.objects.select_related(
            'content_type', 
            'processed_by'
        ).prefetch_related('content_object')
        
        filters = {}
        
        if self.request.GET.get('item_name'):
            vet_ct = ContentType.objects.get_for_model(VetSupply)
            office_ct = ContentType.objects.get_for_model(OfficeSupply)
            
            vet_ids = VetSupply.objects.filter(
                name__icontains=self.request.GET['item_name']
            ).values_list('id', flat=True)
            office_ids = OfficeSupply.objects.filter(
                name__icontains=self.request.GET['item_name']
            ).values_list('id', flat=True)
            
            queryset = queryset.filter(
                (Q(content_type=vet_ct) & Q(object_id__in=vet_ids)) |
                (Q(content_type=office_ct) & Q(object_id__in=office_ids))
            )

        # Filter by supply type
        supply_type = self.request.GET.get('supply_type')
        if supply_type:
            if supply_type == 'vet':
                ct = ContentType.objects.get_for_model(VetSupply)
            elif supply_type == 'office':
                ct = ContentType.objects.get_for_model(OfficeSupply)
            if supply_type in ['vet', 'office']:
                queryset = queryset.filter(content_type=ct)
        
        # Filter by movement type
        movement_type = self.request.GET.get('movement_type')
        if movement_type:
            filters['movement_type'] = movement_type
            
        if self.request.GET.get('date_from'):
            filters['timestamp__date__gte'] = self.request.GET['date_from']
            
        if self.request.GET.get('date_to'):
            filters['timestamp__date__lte'] = self.request.GET['date_to']
        
        return queryset.filter(**filters).order_by('-timestamp')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['from_type'] = self.request.GET.get('from_type', '')
        
        # Add supply types
        context['supply_types'] = [
            ('vet', 'Veterinary'),
            ('office', 'Office')
        ]
        context['selected_supply_type'] = self.request.GET.get('supply_type', '')
        
        # Add movement types
        context['movement_types'] = InventoryMovement.MOVEMENT_TYPES
        context['selected_type'] = self.request.GET.get('movement_type', '')
        
        return context

# ...existing code...

@login_required
def export_transactions(request):
    view = TransactionHistoryView()
    view.request = request
    queryset = view.get_queryset()
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="transactions.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['Date', 'Supply Type', 'Item', 'Type', 'Quantity', 'Previous Qty', 'New Qty', 'Processed By', 'Notes'])
    
    for t in queryset:
        item = t.content_object
        supply_type = 'Veterinary' if isinstance(item, VetSupply) else 'Office'
        item_name = item.name if item else 'Unknown Item'
        
        writer.writerow([
            t.timestamp.strftime('%Y-%m-%d %H:%M'),
            supply_type,
            item_name,
            dict(InventoryMovement.MOVEMENT_TYPES)[t.movement_type],
            t.quantity,
            t.previous_quantity,
            t.current_quantity,
            t.processed_by.username if t.processed_by else 'System',
            t.notes
        ])
    
    return response

#...existing code...

@login_required
def adjust_inventory(request):
    if request.method == 'POST':
        supply_id = request.POST.get('supply_id')
        adjustment = int(request.POST.get('adjustment', 0))
        reason = request.POST.get('reason', '')
        
        try:
            supply = VetSupply.objects.get(id=supply_id)
            previous_quantity = supply.quantity
            new_quantity = previous_quantity + adjustment
            
            if new_quantity < 0:
                messages.error(request, "Adjustment would result in negative stock")
                return redirect('vet_supplies:supply-list')
            
            supply.quantity = new_quantity
            supply.save()
            
            InventoryMovement.objects.create(
                content_type=ContentType.objects.get_for_model(supply),
                object_id=supply.id,
                movement_type='ADJ',
                quantity=abs(adjustment),
                processed_by=request.user,
                previous_quantity=previous_quantity,
                current_quantity=new_quantity,
                notes=f"Manual adjustment: {reason}. Change: {'+'if adjustment > 0 else ''}{adjustment}"
            )
            
            messages.success(request, f'Successfully adjusted {supply.name} quantity')
            
        except VetSupply.DoesNotExist:
            messages.error(request, "Supply not found")
        except Exception as e:
            messages.error(request, f"Error adjusting inventory: {str(e)}")
            
    return redirect('vet_supplies:supply-list')

# ...existing code...

class SupplyUpdate(LoginRequiredMixin, UpdateView):
    model = VetSupply
    form_class = VetSupplyForm
    template_name = 'vet_supplies/supply_form.html'
    success_url = reverse_lazy('vet_supplies:supply-list')

    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(self.request, f'{self.object.name} updated successfully')
        return response

# Add transaction recording for delete operations
def delete_supply(request, pk):
    try:
        supply = VetSupply.objects.get(pk=pk)
        InventoryMovement.objects.create(
            content_type=ContentType.objects.get_for_model(supply),
            object_id=supply.id,
            movement_type='DEL',
            quantity=supply.quantity,
            processed_by=request.user,
            previous_quantity=supply.quantity,
            current_quantity=0,
            notes=f"Deleted item: {supply.name}"
        )
        supply.delete()
        messages.success(request, 'Supply deleted successfully')
    except Exception as e:
        messages.error(request, f'Error deleting supply: {str(e)}')
    return redirect('vet_supplies:supply-list')
