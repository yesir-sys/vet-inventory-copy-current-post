from django.contrib import messages
from django.shortcuts import render, HttpResponse, redirect
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views.generic import ListView
from django.contrib.contenttypes.models import ContentType
from django.utils import timezone
from django.db.models import Count, Sum, Q, F
from django.db.models.functions import TruncDay, TruncWeek, TruncMonth
from datetime import timedelta, datetime
import json
from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from io import BytesIO
from .models import InventoryMovement, StockLevel
from vet_supplies.models import VetSupply
from office_supplies.models import OfficeSupply

@login_required
def dashboard(request):
    try:
        # Log the access attempt
        print(f"User {request.user} accessing dashboard")
        days = int(request.GET.get('days', 30))
        end_date = timezone.now()
        start_date = end_date - timedelta(days=days)
        timeframe = request.GET.get('timeframe', 'daily')
        active_tab = request.GET.get('active_tab', 'vet')  # Get active tab from URL
        
        # Vet Supplies Statistics
        vet_total = VetSupply.objects.count()
        vet_low_stock = VetSupply.objects.filter(quantity__lte=F('reorder_level')).count()
        vet_out_of_stock = VetSupply.objects.filter(quantity=0).count()
        vet_normal_stock = vet_total - vet_low_stock - vet_out_of_stock
        
        # Office Supplies Statistics
        office_total = OfficeSupply.objects.count()
        office_low_stock = OfficeSupply.objects.filter(quantity__lte=F('reorder_level')).count()
        office_out_of_stock = OfficeSupply.objects.filter(quantity=0).count()
        office_normal_stock = office_total - office_low_stock - office_out_of_stock
        
        # Get trend data based on selected timeframe
        trunc_func = {
            'daily': TruncDay,
            'weekly': TruncWeek,
            'monthly': TruncMonth
        }.get(timeframe, TruncDay)

        # Get movements by type with proper date truncation
        vet_movements = InventoryMovement.objects.filter(
            timestamp__range=(start_date, end_date),
            content_type__model='vetsupply'
        ).annotate(
            date=trunc_func('timestamp')
        ).values('date', 'movement_type').annotate(
            total=Sum('quantity')
        ).order_by('date')

        office_movements = InventoryMovement.objects.filter(
            timestamp__range=(start_date, end_date),
            content_type__model='officesupply'
        ).annotate(
            date=trunc_func('timestamp')
        ).values('date', 'movement_type').annotate(
            total=Sum('quantity')
        ).order_by('date')

        # Movement counts
        vet_in_count = vet_movements.filter(movement_type='IN').count()
        vet_out_count = vet_movements.filter(movement_type='OUT').count()
        office_in_count = office_movements.filter(movement_type='IN').count()
        office_out_count = office_movements.filter(movement_type='OUT').count()
        
        context = {
            'vet_stats': {
                'total': vet_total,
                'low_stock': vet_low_stock,
                'out_of_stock': vet_out_of_stock,
                'normal_stock': vet_normal_stock,
                'stock_in': vet_in_count,
                'stock_out': vet_out_count,
            },
            'office_stats': {
                'total': office_total,
                'low_stock': office_low_stock,
                'out_of_stock': office_out_of_stock,
                'normal_stock': office_normal_stock,
                'stock_in': office_in_count,
                'stock_out': office_out_count,
            },
            'selected_days': days,
        }

        # Prepare chart data
        trend_data = {
            'vet': format_trend_data(vet_movements, timeframe),
            'office': format_trend_data(office_movements, timeframe)
        }
        
        context.update({
            'trend_data': json.dumps(trend_data),
            'selected_timeframe': timeframe,
            'active_tab': active_tab,  # Add active tab to context
            'from_type': request.GET.get('from_type', '')  # Add this line
        })
        
        return render(request, 'reports/dashboard.html', context)
    except Exception as e:
        messages.error(request, f"Error accessing reports: {str(e)}")
        return redirect('home')

def format_trend_data(movements, timeframe):
    date_format = {
        'daily': '%Y-%m-%d',
        'weekly': '%Y-W%W',
        'monthly': '%Y-%m'
    }.get(timeframe, '%Y-%m-%d')

    # Get distribution data first
    normal_count = movements.filter(movement_type='IN').count()
    low_count = movements.filter(movement_type='OUT').count()
    out_count = movements.filter(movement_type='ADJ').count()

    # Get all dates in range
    dates = sorted(set(m['date'].strftime(date_format) for m in movements))
    in_data = {d: 0 for d in dates}
    out_data = {d: 0 for d in dates}

    for m in movements:
        date_str = m['date'].strftime(date_format)
        if m['movement_type'] == 'IN':
            in_data[date_str] = m.get('total', 0) or 0
        elif m['movement_type'] == 'OUT':
            out_data[date_str] = m.get('total', 0) or 0

    # Format dates for display
    formatted_dates = []
    for date_str in dates:
        if timeframe == 'weekly':
            year, week = date_str.split('-W')
            formatted_dates.append(f'Week {week}')
        elif timeframe == 'monthly':
            formatted_dates.append(datetime.strptime(date_str, '%Y-%m').strftime('%b %Y'))
        else:
            formatted_dates.append(datetime.strptime(date_str, '%Y-%m-%d').strftime('%d %b'))

    # Get distribution data for pie chart
    normal_count = movements.filter(movement_type='IN').count()  # Items with normal stock levels
    low_count = movements.filter(movement_type='OUT').count()    # Items with low stock
    out_count = movements.filter(movement_type='ADJ').count()    # Items out of stock

    return {
        'dates': formatted_dates,
        'in_data': list(in_data.values()),
        'out_data': list(out_data.values()),
        'distribution': {
            'normal': normal_count,
            'low': low_count,
            'out': out_count
        }
    }

@login_required
def export_report(request):
    try:
        report_type = request.GET.get('type', 'vet')
        export_format = request.GET.get('format', 'excel')
        report_detail = request.GET.get('report_type', 'detailed')
        
        # Get correct queryset based on type
        if report_type == 'vet':
            queryset = VetSupply.objects.all().select_related('category')
        else:
            queryset = OfficeSupply.objects.all().select_related('category')

        if export_format == 'excel':
            if report_detail == 'detailed':
                response = generate_detailed_excel(queryset, report_type)
            else:
                response = generate_summary_excel(queryset, report_type)
            return response
        elif export_format == 'pdf':
            return HttpResponse("PDF export coming soon", status=501)
        
        return HttpResponse("Invalid format", status=400)
        
    except Exception as e:
        print(f"Export error: {str(e)}")
        return HttpResponse(f"Export failed: {str(e)}", status=500)

def generate_detailed_excel(queryset, report_type):
    wb = Workbook()
    ws = wb.active
    ws.title = f"{report_type.title()} Inventory"
    
    # Add title and metadata
    ws.merge_cells('A1:G1')
    ws['A1'] = f"{report_type.title()} Detailed Inventory Report"
    ws['A1'].font = Font(size=16, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws['A2'] = f"Generated on: {timezone.now().strftime('%Y-%m-%d %H:%M')}"
    ws['A3'] = f"Total Items: {queryset.count()}"
    
    # Headers with styling
    headers = ['Item Name', 'Category', 'Current Stock', 'Reorder Level', 
              'Status', 'Last Updated', 'Expiration Date']
    
    header_style = NamedStyle(name='header')
    header_style.font = Font(bold=True, color='FFFFFF')
    header_style.fill = PatternFill(start_color='2B7C85', end_color='2B7C85', fill_type='solid')
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col)
        cell.value = header
        cell.style = header_style
    
    # Data rows
    for row, item in enumerate(queryset, 6):
        ws.cell(row=row, column=1, value=item.name)
        ws.cell(row=row, column=2, value=item.category.name)
        ws.cell(row=row, column=3, value=item.quantity)
        ws.cell(row=row, column=4, value=item.reorder_level)
        
        # Status cell with conditional formatting
        status_cell = ws.cell(row=row, column=5)
        if item.quantity <= item.reorder_level:
            status_cell.value = 'Low Stock'
            status_cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
        else:
            status_cell.value = 'OK'
            status_cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
            
        ws.cell(row=row, column=6, value=item.last_updated.strftime('%Y-%m-%d'))
        ws.cell(row=row, column=7, value=item.expiration_date.strftime('%Y-%m-%d') if item.expiration_date else 'N/A')
    
    # Auto-adjust columns
    for column in ws.columns:
        max_length = 0
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{report_type}_detailed_inventory.xlsx"'
    wb.save(response)
    return response

def generate_summary_excel(queryset, report_type):
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    
    # Add title
    ws['A1'] = f"{report_type.title()} Inventory Summary"
    ws['A1'].font = Font(size=16, bold=True)
    
    # Summary statistics
    summary_data = [
        ('Total Items', queryset.count()),
        ('Low Stock Items', queryset.filter(quantity__lte=F('reorder_level')).count()),
        ('Out of Stock', queryset.filter(quantity=0).count()),
        ('Expiring Soon', queryset.filter(
            expiration_date__lte=timezone.now().date() + timedelta(days=30),
            expiration_date__gt=timezone.now().date()
        ).count()),
    ]
    
    for row, (label, value) in enumerate(summary_data, 3):
        ws[f'A{row}'] = label
        ws[f'B{row}'] = value
        ws[f'A{row}'].font = Font(bold=True)
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{report_type}_summary_inventory.xlsx"'
    wb.save(response)
    return response

@login_required 
def view_by_category(request, supply_type, category):
    if supply_type == 'vet':
        return redirect('vet_supplies:supply-list', category=category)
    return redirect('office_supplies:supply-list', category=category)

@login_required
def view_low_stock(request, supply_type):
    if supply_type == 'vet':
        return redirect('vet_supplies:low-stock')
    return redirect('office_supplies:low-stock')

@login_required
def view_movements(request, supply_type, movement_type):
    # Add movement type to session for filtering
    request.session['movement_filter'] = movement_type
    if supply_type == 'vet':
        return redirect('vet_supplies:supply-list')
    return redirect('office_supplies:supply-list')
